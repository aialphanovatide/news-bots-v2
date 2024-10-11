
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from pathlib import Path
import tempfile
import re

def clean_text(text):
    """
    Clean and normalize the input text.

    This function performs the following operations:
    1. Converts the text to lowercase.
    2. Removes all special characters, keeping only alphanumeric characters and spaces.
    3. Replaces multiple consecutive spaces with a single space.
    4. Strips leading and trailing whitespace.

    Args:
        text (str): The input text to be cleaned.

    Returns:
        str: The cleaned and normalized text.
    """
    # Convert to lowercase and remove special characters
    cleaned = re.sub(r'[^a-zA-Z0-9\s]', '', text.lower())
    # Replace multiple spaces with a single space
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned

def extract_excel_content(file_path):
    """
    Extract keywords and blacklist content from an Excel file.

    This function reads an Excel file and extracts keywords and blacklist items
    from sheets with names containing 'keyword' or 'blacklist' respectively.

    Args:
        file_path (str): The path to the Excel file.

    Returns:
        tuple: A tuple containing two strings:
            - keywords_str: A comma-separated string of extracted keywords.
            - blacklist_str: A comma-separated string of extracted blacklist items.

    Raises:
        ValueError: If no 'keywords' or 'blacklist' data is found in the Excel file.
        Exception: For any other errors that occur during file processing.

    Notes:
        - The function uses case-insensitive matching for sheet names and column headers.
        - Keywords and blacklist items are cleaned and converted to lowercase.
        - Duplicate entries are automatically removed as sets are used for storage.
        - The workbook is always closed after processing, even if an exception occurs.
    """
    try:
        wb = load_workbook(filename=file_path, read_only=True)
        
        keywords = set()  # Use set to automatically remove duplicates
        blacklist = set()  # Use set to automatically remove duplicates

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            first_row = [str(cell.value).lower() if cell.value else '' for cell in next(sheet.iter_rows())]

            if 'keyword' in sheet_name.lower():
                keywords_col_index = next((i for i, col in enumerate(first_row) if 'keyword' in col), None)
                
                if keywords_col_index is not None:
                    keywords.update(clean_text(str(row[keywords_col_index].value).strip().lower()) for row in sheet.iter_rows(min_row=2) if row[keywords_col_index].value)
            
            elif 'blacklist' in sheet_name.lower():
                blacklist_col_index = next((i for i, col in enumerate(first_row) if 'blacklist' in col), None)
                
                if blacklist_col_index is not None:
                    blacklist.update(clean_text(str(row[blacklist_col_index].value).strip().lower()) for row in sheet.iter_rows(min_row=2) if row[blacklist_col_index].value)

        if not keywords and not blacklist:
            raise ValueError("No 'keywords' or 'blacklist' data found in the Excel file")

        # Remove any empty strings and join with commas
        keywords_str = ','.join(sorted(filter(None, keywords)))
        blacklist_str = ','.join(sorted(filter(None, blacklist)))

        return keywords_str, blacklist_str
    except Exception as e:
        raise
    finally:
        wb.close()
        
def process_uploaded_file(file):
    """
    Process an uploaded Excel file and extract its content.

    This function handles the upload of an Excel file, saves it temporarily,
    extracts its content, and then removes the temporary file.

    Args:
        file: A file object representing the uploaded Excel file.

    Returns:
        tuple: A tuple containing two strings:
            - keywords_str: A comma-separated string of extracted keywords.
            - blacklist_str: A comma-separated string of extracted blacklist items.

    Raises:
        ValueError: If no file is selected or if the file type is not supported.
        Exception: Any exception raised during file processing is propagated.

    Notes:
        - Only .xls and .xlsx file extensions are supported.
        - The file is temporarily saved and then deleted after processing.
        - The actual content extraction is performed by the extract_excel_content function.
    """
    if file.filename == '':
        raise ValueError("No file selected")
    
    filename = secure_filename(file.filename)
    file_extension = Path(filename).suffix
    
    if file_extension.lower() not in ['.xls', '.xlsx']:
        raise ValueError("Only Excel files (.xls, .xlsx) are supported")
    
    # Create a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as temp_file:
        file.save(temp_file.name)
        temp_file_path = Path(temp_file.name)

    try:
        data = extract_excel_content(temp_file_path)
        return data
    finally:
        # Ensure the temporary file is deleted
        if temp_file_path.exists():
            temp_file_path.unlink()